import os
import re
import itertools
from html.parser import HTMLParser
from typing import List, Dict, Any
from pypdf import PdfReader
import openpyxl

# Constants
DOCUMENTS_DIR = "documents"
CONTEXT_OFFSET = 40
PROXIMITY_THRESHOLD = 50
ENCODING_UTF8 = 'utf-8'
OP_AND = " AND "
OP_OR = " OR "
EXT_TXT = '.txt'
EXT_HTML = '.html'
EXT_PDF = '.pdf'
EXT_XLSX = '.xlsx'
EXT_XLS = '.xls'


class TextExtractorHTML(HTMLParser):
    """
    Standard library HTML parser to avoid using BeautifulSoup.
    Accumulates text data from tags.
    """
    def __init__(self):
        super().__init__()
        self.text_parts = []

    def handle_data(self, data):
        """ Capture text content within tags. """
        clean_text = data.strip()
        if clean_text:
            self.text_parts.append(clean_text)

    def get_text(self) -> str:
        """ Return the joined text content. """
        return "\n".join(self.text_parts)


def get_context_snippet(text: str, match_index: int) -> str:
    """ Return a substring around the match index with ellipsis. """
    start = max(0, match_index - CONTEXT_OFFSET)
    end = min(len(text), match_index + CONTEXT_OFFSET)
    snippet = text[start:end].replace('\n', ' ').strip()
    return f"...{snippet}..."


def check_proximity(text: str, query: str) -> List[int]:
    """
    Check if all terms in the query exist within a close distance.
    Returns a list of start indices for matches.
    """
    terms = [t.strip() for t in query.split()]
    term_indices = []

    # Find all positions for each term
    for term in terms:
        pattern = re.escape(term)
        matches = [m.start() for m in re.finditer(pattern, text, re.IGNORECASE)]
        if not matches:
            return []
        term_indices.append(matches)

    valid_indices = []
    # Check all combinations of term positions
    for combination in itertools.product(*term_indices):
        min_pos = min(combination)
        max_pos = max(combination)
        
        if (max_pos - min_pos) <= PROXIMITY_THRESHOLD:
            valid_indices.append(min_pos)

    return sorted(list(set(valid_indices)))


def search_content(text: str, filename: str, f_type: str, loc: str, query: str, use_regex: bool) -> List[Dict[str, Any]]:
    """ Core search logic handling Regex, Boolean (AND/OR), and Proximity. """
    results = []
    indices = []

    # Boolean Logic
    if not use_regex and (OP_AND in query or OP_OR in query):
        is_and = OP_AND in query
        operator = OP_AND if is_and else OP_OR
        parts = query.split(operator)

        if is_and:
            if all(p in text for p in parts):
                indices.append(text.find(parts[0]))
        else:
            for p in parts:
                if p in text:
                    indices.append(text.find(p))
                    break

    # Proximity (Space separated, no operators, not regex)
    elif not use_regex and " " in query:
        indices = check_proximity(text, query)

    # Regex or Simple Match
    else:
        if use_regex:
            match = re.search(query, text)
            if match:
                indices.append(match.start())
        else:
            if query in text:
                indices.append(text.find(query))

    # Format results
    for idx in indices:
        results.append({
            "file": filename,
            "type": f_type,
            "location": loc,
            "context": get_context_snippet(text, idx)
        })

    return results


def parse_txt(path: str, query: str, regex: bool) -> List[Dict]:
    """ Read and search TXT files line by line. """
    results = []
    try:
        with open(path, 'r', encoding=ENCODING_UTF8, errors='ignore') as f:
            for i, line in enumerate(f, 1):
                if line.strip():
                    results.extend(search_content(
                        line, os.path.basename(path), "txt", f"Line {i}", query, regex
                    ))
    except IOError:
        pass
    return results


def parse_html(path: str, query: str, regex: bool) -> List[Dict]:
    """ Read and search HTML files using standard library parser. """
    results = []
    try:
        with open(path, 'r', encoding=ENCODING_UTF8, errors='ignore') as f:
            parser = TextExtractorHTML()
            parser.feed(f.read())
            lines = parser.get_text().splitlines()
            
            for i, line in enumerate(lines, 1):
                results.extend(search_content(
                    line, os.path.basename(path), "html", f"Line {i}", query, regex
                ))
    except Exception:
        pass
    return results


def parse_pdf(path: str, query: str, regex: bool) -> List[Dict]:
    """ Read and search PDF files. Requires pypdf. """
    results = []
    try:
        reader = PdfReader(path)
        for i, page in enumerate(reader.pages, 1):
            text = page.extract_text() or ""
            for line_num, line in enumerate(text.splitlines(), 1):
                results.extend(search_content(
                    line, os.path.basename(path), "pdf", f"Page {i}, Line {line_num}", query, regex
                ))
    except Exception:
        pass
    return results


def parse_excel(path: str, query: str, regex: bool) -> List[Dict]:
    """ Read Excel files. """
    results = []
    try:
        # data_only=True gets the calculated value instead of the formula
        wb = openpyxl.load_workbook(path, data_only=True)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            for row in sheet.iter_rows():
                for cell in row:
                    # Explicitly check for None to include 0 values
                    if cell.value is not None:
                        text = str(cell.value).strip()
                        
                        # Only search if text is not empty
                        if text:
                            loc = f"{sheet_name} | {cell.coordinate}"
                            results.extend(search_content(
                                text, os.path.basename(path), "xlsx", loc, query, regex
                            ))
    except Exception as e:
        print(f"Error reading Excel file {path}: {e}")
    return results


def process_search(query: str, extensions: List[str], use_regex: bool) -> List[Dict]:
    """ Main entry point. Iterates through the relative documents folder. """
    final_results = []
    
    if not os.path.exists(DOCUMENTS_DIR):
        return []

    for filename in os.listdir(DOCUMENTS_DIR):
        filepath = os.path.join(DOCUMENTS_DIR, filename)
        
        if not os.path.isfile(filepath):
            continue

        _, ext = os.path.splitext(filename)
        ext = ext.lower()

        if extensions and ext not in extensions:
            continue

        if ext == EXT_TXT:
            final_results.extend(parse_txt(filepath, query, use_regex))
        elif ext == EXT_HTML:
            final_results.extend(parse_html(filepath, query, use_regex))
        elif ext == EXT_PDF:
            final_results.extend(parse_pdf(filepath, query, use_regex))
        elif ext in [EXT_XLSX, EXT_XLS]:
            final_results.extend(parse_excel(filepath, query, use_regex))

    return final_results